import time
import configparser

from selenium import webdriver
from openpyxl import load_workbook
from jinja2 import Template


class EmailWorker(object):
    def __init__(self):
        conf = configparser.ConfigParser()
        conf.read("config.ini")
        self.work_email = conf.get("account", "work_email")
        self.work_password = conf.get("account", "work_password")
        self.email_theme = conf.get("theme", "email_theme")
        self.chromedriver_path = conf.get("chromedriver", "chromedriver_path")
        self.browser = webdriver.Chrome(executable_path=self.chromedriver_path)
        self.excel_path = conf.get("excel", "excel_path")
        self.login_url = "https://qiye.163.com/login/"
        self.template = """
            <!DOCTYPE html>
            <html lang="en">
            
            <body>
                <table border="1" width="100%" cellspacing="0" cellpadding="2" bordercolor="#000000"  border-collapse="collapse">
                <tr style="background:#fff;" >
                {% for item in items %}
                    <td style="empty-cells:show" >{{ item }}</td>
                {% endfor %}    
                  
                </tr>   
                <tr style="background:#fafafa;">
                {% for salary in salarys %}
                  <td style="empty-cells:show" >{{ salary }}</td>
                {% endfor %}
                </tr>
                </table>
                
            </body>
            </html>
        """

    def login_to_write(self):
        self.browser.get(self.login_url)
        self.browser.find_element_by_id("accname").send_keys(self.work_email)
        self.browser.find_element_by_id("accpwd").send_keys(self.work_password)
        self.browser.find_element_by_css_selector(".loginbtn .js-loginbtn").click()
        time.sleep(2)
        # write email button
        self.browser.find_element_by_css_selector("#_mail_component_37_37 .pd0").click()

    def ready_write(self):
        # more choices button
        self.browser.find_element_by_partial_link_text("更多发送选项").click()
        time.sleep(1)
        # encrypt choice
        self.browser.find_element_by_xpath("//div[@class='ht0']/span[5]").click()
        
        # html code choice
        try:
            self.browser.find_element_by_xpath('//a[@title="切换到全部功能"]').click()
        except:
            pass
        time.sleep(1)
        self.browser.find_element_by_class_name('APP-editor-btn-switchSource').click()
        

    def do_write(self):
        result = self.generate_email()
        while True:
            
            try:
                # recipients
                email, encrypt, template = next(result)
                # print(email, encrypt)
                self.browser.find_element_by_class_name("nui-editableAddr-ipt").send_keys(email)
                # theme
                self.browser.find_element_by_css_selector(".bF0 .nui-ipt .nui-ipt-input").send_keys(self.email_theme)
                # email body
                self.browser.find_element_by_css_selector(".APP-editor-textarea").send_keys(template)
                # encrypt number
                self.browser.find_element_by_css_selector(".cA0 .nui-ipt-input").send_keys(encrypt)

                self.browser.find_element_by_css_selector(".jA0 .js-component-button").click()
                time.sleep(2)

                # continue to write
                self.browser.find_element_by_xpath("//section[@class='sp0']/div[3]/a[3]").click()
                time.sleep(1)
                self.ready_write()
            except Exception as e:
                with open('log.txt', 'w') as f:
                    f.write(e)
                break
        # close the window when finished  work
        self.browser.close()
    
    def generate_email(self):
        wb = load_workbook(self.excel_path, data_only=True)
        sheet = wb.get_sheet_by_name("Sheet1")
        # pop encrypt and email
        items = [i.value for i in sheet["1"]][2:]
        
        for x in range(2, sheet.max_row+1):
            salarys = [y.value if y.value is not None else "&nbsp;" for y in sheet["{}".format(x)] ]
            encrypt = salarys.pop(0)
            email = salarys.pop(0)
            context = {"items": items, "salarys": salarys}
            template = Template(self.template)
            yield email, encrypt, template.render(**context)


if __name__ == '__main__':
    work = EmailWorker()
    work.login_to_write()
    work.ready_write()
    work.do_write()
